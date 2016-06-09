# Created by Michael Scales
# Version 2.0
# This program will take user modified excel spreadsheets 
# parses the chronicall reports, collates the information and generates a mostly completed Daily SLA Report
# Jan 11 2016

# -*- coding: utf-8 -*-
import datetime
import subprocess
import sys
import traceback
from os import path

import openpyxl
from openpyxl.compat import range

from src import (Client)
from src.CONSTANTS import (CALL_DETAILS_FIRST_PAGE_NAME,
                           ARG_PATH,
                           EXC_PATH,
                           SELF_PATH,
                           MOD_CALL_DETAILS_PATH)
from src.email_reader import (get_voice_mail, get_downloads)
from src.extractor import (find_duplicates_abandon_group_report, process_hunt_group_report, process_abandon_calls,
                           process_call_details_report, parse_chronicall_report)
from src.input_valididation import valid_input
from src.utils import (initialize_reports, process_duplicates, remove_voice_mail_calls, white_wash,
                       prepare_raw_call_details, check_date, get_client_dictionary, process_report)

# !/usr/bin/env python3
# Add the executing directory to the class path
sys.path.append(path.dirname(path.dirname(path.abspath(__file__))))
print("Daily SLA Program... Created by Michael Scales")


def calc_data(hunt_grp, abandon_grp, details,
              client, voice_mail_dict, client_dict):
    client_name = client.get_name()
    client_name_verbose = client_dict[client_name]

    lost_calls, avg_wait_lost_call = process_abandon_calls(abandon_grp, client_name)

    client.set_abandon_grp(lost_calls, avg_wait_lost_call)
    # Everything from Abandon Calls report

    number_of_voice_mails, voice_mail_to_remove = remove_voice_mail_calls(abandon_grp,
                                                                          voice_mail_dict,
                                                                          client_name_verbose,
                                                                          client_dict)
    client.set_voice_mails(number_of_voice_mails, voice_mail_to_remove)

    (hunt_group_calls_answered,
     hunt_group_avg_call_duration,
     hunt_group_avg_wait_answered_call,
     hunt_group_longest_answered,
     hunt_group_calls_less_than_15sec,
     hunt_group_calls_less_than_30sec,
     hunt_group_calls_less_than_45sec,
     hunt_group_calls_less_than_60sec,
     hunt_group_calls_greater_than_g60sec) = process_hunt_group_report(hunt_grp)
    client.set_hunt_grp(hunt_group_calls_answered,
                        hunt_group_avg_call_duration,
                        hunt_group_avg_wait_answered_call,
                        hunt_group_longest_answered)
    # Sets values for object

    client.set_ticker(hunt_group_calls_less_than_15sec,
                      hunt_group_calls_less_than_30sec,
                      hunt_group_calls_less_than_45sec,
                      hunt_group_calls_less_than_60sec,
                      hunt_group_calls_greater_than_g60sec)  # Sets values for object

    # combines call details information with huntGRP and abandonGRP
    (call_details_calls_answered,
     call_details_call_duration,
     call_details_wait_answered,
     call_details_longest_hold_ans,
     call_details_calls_less_than_15sec,
     call_details_calls_less_than_30sec,
     call_details_calls_less_than_45sec,
     call_details_calls_less_than_60sec,
     call_details_calls_greater_than_60sec) = process_call_details_report(details, client_name, client.get_24hr())
    # Everything from hunt group reports

    if call_details_calls_answered == 0:
        pass
    else:
        hunt_group_total_duration_wait_answered = hunt_group_avg_wait_answered_call * hunt_group_calls_answered
        hunt_group_total_call_duration = hunt_group_avg_call_duration * hunt_group_calls_answered

        combined_duration = hunt_group_total_call_duration + call_details_call_duration
        combined_calls = hunt_group_calls_answered + call_details_calls_answered

        client.set_call_details_calls(combined_calls)

        client.add_ticker(call_details_calls_less_than_15sec,
                          call_details_calls_less_than_30sec,
                          call_details_calls_less_than_45sec,
                          call_details_calls_less_than_60sec,
                          call_details_calls_greater_than_60sec)

        combined_avg_duration = combined_duration / combined_calls
        combined_wait_answered_duration = hunt_group_total_duration_wait_answered + call_details_wait_answered
        combined_avg_wait_answered = combined_wait_answered_duration / combined_calls
        client.set_call_duration(combined_avg_duration, combined_avg_wait_answered)

        if call_details_longest_hold_ans >= hunt_group_longest_answered:
            client.set_longest_wait_answered(call_details_longest_hold_ans)


def create_client_report(client_list_info):
    client_report = []
    hunt_groups = []
    for row_number in range(1, client_list_info.max_row + 1):
        client = Client(valid_input('A', row_number, client_list_info, 'N'))
        client.set_24hr(valid_input('C', row_number, client_list_info, 'N'))
        hunt_group_path = valid_input('B', row_number, client_list_info, 'S')
        client_report.append(client)
        hunt_groups.append(hunt_group_path)
    return client_report, hunt_groups


def run_report(abandon_group_file, new_call_details_file, report_date_datetime, client_report,
               hunt_groups, voice_mail_dict, client_dict):
    print("Attempting to collate data.")
    date_string = report_date_datetime.strftime("%m%d%Y")
    for index, client in enumerate(client_report):
        hunt = openpyxl.load_workbook(SELF_PATH + (hunt_groups[index]))
        check_date(hunt, date_string, 'hunt')
        calc_data(hunt,
                  abandon_group_file,
                  new_call_details_file,
                  client,
                  voice_mail_dict,
                  client_dict)


def main(report_date_datetime):
    try:
        white_wash()
        client_dict = get_client_dictionary()
        # Dictionary of client names/numbers
        # Pre-condition: correct data in client_dict.xlsx
        # Post-condition: Returns a dictionary for converting between client number and client name
        # Ex. Client Number (int) : Client Name (string)
        print("report date: %s" % report_date_datetime)
        get_downloads(report_date_datetime)
        # Downloads hunt group reports, call details, cradle 2 grave and abandon group files
        # Pre-condition: None
        # Post-condition: Saves contents to Archives
        get_downloads(report_date_datetime, 'raw')
        # Downloads hunt group reports, call details, cradle 2 grave and abandon group files
        # Pre-condition: None
        # Post-condition: Saves contents to \raw\ for program use
        subprocess.call([EXC_PATH, ARG_PATH], shell=False)
        # Executes Microsoft Office Conversion Software
        # Pre-condition: None
        # Post-condition: Any .xls files located in \raw\ are converted to .xlsx. Files are saved to:
        # \converted_files_come_here\Desktop\development\Daily SLA Parser - Automated Version\raw
        (client_list_info,
         call_details_file,
         abandon_group_file,
         cradle_to_grave_file) = initialize_reports(report_date_datetime)
        # Opens and stores Call Details.xlsx, Cradle 2 Grave.xlsx and Abandon Group Reports.xlsx
        # Pre-condition: The file date matches the report_date
        # Post-condition: Returns file variables and date for use in program execution
        voice_mail_dict = get_voice_mail(report_date_datetime)
        # Creates a dictionary collection of voicemails received for the report_date
        # Pre-condition: None
        # Post-condition: Returns a default_dict(list) which stores the client, last 4 of the phone number,
        # call time. Ex. Key: AAP Value List: ['9063 + 09:29:25', '3619 + 13:21:14', '2071 + 16:10:25']
        ws = call_details_file.get_sheet_by_name(CALL_DETAILS_FIRST_PAGE_NAME)
        # Opens Call Details Workbook as a Worksheet
        # Pre-condition: None
        # Post-condition: Stores a worksheet which can be iterated through
        call_id_list = parse_chronicall_report(cradle_to_grave_file)
        # List of Call ID's to be used for calculations. Program generated report "Call Details"
        # Pre-condition: None
        # Post-condition: Returns a default_dict(list) which stores the client, last 4 of the phone number,
        # + call time. Ex. Key: AAP Value List: ['9063 + 09:29:25', '3619 + 13:21:14', '2071 + 16:10:25']
        prepare_raw_call_details(ws, call_details_file, call_id_list)
        # List of Call ID's to be used for calculations. Program generated report "Call Details"
        # Pre-condition: None
        # Post-condition: Returns a default_dict(list) which stores the client, last 4 of the phone number,
        # + call time. Ex. Key: AAP Value List: ['9063 + 09:29:25', '3619 + 13:21:14', '2071 + 16:10:25']
        client_report, hunt_groups = create_client_report(client_list_info)
        # List of Call ID's to be used for calculations. Program generated report "Call Details"
        # Pre-condition: None
        # Post-condition: Returns a default_dict(list) which stores the client, last 4 of the phone number,
        # + call time. Ex. Key: AAP Value List: ['9063 + 09:29:25', '3619 + 13:21:14', '2071 + 16:10:25']
        new_call_details_file = openpyxl.load_workbook(SELF_PATH + MOD_CALL_DETAILS_PATH)
        # List of Call ID's to be used for calculations. Program generated report "Call Details"
        # Pre-condition: None
        # Post-condition: Returns a default_dict(list) which stores the client, last 4 of the phone number,
        # + call time. Ex. Key: AAP Value List: ['9063 + 09:29:25', '3619 + 13:21:14', '2071 + 16:10:25']
        run_report(abandon_group_file, new_call_details_file, report_date_datetime,
                   client_report, hunt_groups, voice_mail_dict, client_dict)
        # List of Call ID's to be used for calculations. Program generated report "Call Details"
        # Pre-condition: None
        # Post-condition: Returns a default_dict(list) which stores the client, last 4 of the phone number,
        # + call time. Ex. Key: AAP Value List: ['9063 + 09:29:25', '3619 + 13:21:14', '2071 + 16:10:25']
        final_duplicate_dict = find_duplicates_abandon_group_report(abandon_group_file)
        # List of Call ID's to be used for calculations. Program generated report "Call Details"
        # Pre-condition: None
        # Post-condition: Returns a default_dict(list) which stores the client, last 4 of the phone number,
        # + call time. Ex. Key: AAP Value List: ['9063 + 09:29:25', '3619 + 13:21:14', '2071 + 16:10:25']
        process_duplicates(final_duplicate_dict, client_report)
        # Removes duplicate abandoned calls. (Same Tel. # within 1 Hr)
        # Pre-condition: None
        # Post-condition: Duplicate calls are removed from the data for the final report.
        process_report(client_report, abandon_group_file, new_call_details_file,
                       report_date_datetime, voice_mail_dict)
        # Produces the spreadsheet with all collated data.
        # Pre-condition: None
        # Post-condition: An archive of the files used is saved into the \Archive\[This date]
        # Output file is saved to \Output
        white_wash(client_report)
        # Clears folders of all files
    except (SystemExit, KeyboardInterrupt, UserWarning):  # Passes any valid system exit command
        pass
    except Exception as err:
        now = datetime.datetime.now()
        error = traceback.format_exc()
        error_log = open(SELF_PATH + '\\error_logs\\%s_error_log.txt' % report_date_datetime.strftime("%m%d%Y"), 'a')
        error_log.write("\n%r Exception Occurred: %s\n File Date: %r\n%s " % (now.strftime("%m/%d/%Y %H:%M:%S"),
                                                                              err,
                                                                              report_date_datetime.strftime("%m/%d/%Y"),
                                                                              error))
        error_log.close()
    finally:
        print("Packing up the tools to quit.")  # Just cause.


if __name__ == "__main__":
    main(datetime.datetime.today() - datetime.timedelta(days=1))
else:
    from os import sys, path

    sys.path.append(path.dirname(path.dirname(path.abspath(__file__))))
    number_of_runs = int(input("How many days?"))
    while number_of_runs > 0:
        try:
            date_text = input("Month") + input("Day") + input("Year")
            report_date = datetime.datetime.strptime(date_text, '%m%d%Y')
        except ValueError:
            raise ValueError("Incorrect data format, should be %m%d%Y")
        else:
            main(report_date)
        finally:
            number_of_runs -= 1
