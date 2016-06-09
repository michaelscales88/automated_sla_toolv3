# Conversion functions to handle the timestamps that are pulled from excel
# Get seconds from timeString HH:MM:SS
# Contains additional utility functions
import os
import openpyxl
from collections import defaultdict
from src.CONSTANTS import (CALL_DETAILS_FIRST_PAGE_NAME,
                           ABANDON_CALLS_GROUP_PAGE_NAME,
                           SUMMARY_PAGE,
                           SELF_PATH,
                           TEMPLATE_LOCATION)
from .abandongrp import AbandonGrp
from .outfile import (write_outfile)
from .input_valididation import valid_input


def remove_voice_mail_calls(abandon_grp, voice_mail_dict, client, client_dict):
    number_of_voice_mails = len(voice_mail_dict[client])
    voice_mail_to_remove = 0
    voice_mail_by_client_list = voice_mail_dict[client]
    if not voice_mail_by_client_list:
        pass
    else:
        ws = abandon_grp.get_sheet_by_name(ABANDON_CALLS_GROUP_PAGE_NAME)
        for index in range(number_of_voice_mails):
            vm_telephone_number, voice_mail_time = (voice_mail_dict[client][index]).split(" + ")
            for row_number in range(5, ws.max_row + 1):
                lost_call_calling_number = valid_input('F', row_number, ws, 'S')[-4:]
                calling_time = valid_input('I', row_number, ws, 'S')[-8:]
                calling_duration = valid_input('K', row_number, ws, 'N')
                answered = valid_input('G', row_number, ws, 'B')
                abandon_grp_client_name = valid_input('E', row_number, ws, 'N')
                difference = get_sec(voice_mail_time) - get_sec(calling_time)
                is_same_client = (client_dict[abandon_grp_client_name] == client)

                if (vm_telephone_number == lost_call_calling_number
                    and difference < 3600
                    and calling_duration > 19
                    and not answered
                    and is_same_client):
                    voice_mail_to_remove += 1
    return number_of_voice_mails, voice_mail_to_remove


def white_wash(client_report=None):  # Clears used data reports after a final report is output
    if client_report is None:
        client_report = []
    print("Clearing data cache.")
    clear_used_data('\\raw\\')
    clear_used_data('\\converted_files_come_here\\Desktop\\development\\Daily SLA Parser - Automated Version\\raw\\')
    if client_report:
        for client in client_report:
            client.reset()


def check_value_in_range(min_range, max_range, work_sheet, max_or_min):
    return_row = 999
    if max_or_min == "Max":
        step_size = -1
    else:
        step_size = 1
    for row_number in range(min_range, max_range, step_size):
        working_cell = valid_input('A', row_number, work_sheet, 'S')
        working_cell = working_cell.split(' ')
        if working_cell[0].isdigit():
            return_row = row_number
            break
        else:
            pass
    return return_row


def find_min_max_rows(work_sheet):
    raw_max_row = work_sheet.max_row + 1
    min_row = check_value_in_range(1, raw_max_row, work_sheet, "Min")
    max_row = check_value_in_range(raw_max_row, 1, work_sheet, "Max")
    return min_row, max_row


def make_row_dictionary(work_sheet):
    return_dict = {}
    min_row, max_row = find_min_max_rows(work_sheet)
    for row_number in range(min_row, max_row + 1):
        cell_string = valid_input('A', row_number, work_sheet, 'S')
        try:
            split_cell = cell_string.split(' ')
        except ValueError:
            pass
        else:
            if split_cell[0].isdigit() and work_sheet.title == "REPORT":
                return_dict[split_cell[0]] = row_number + 1
            elif split_cell[0].isdigit() and work_sheet.title == "data":
                return_dict[split_cell[0]] = row_number + 2
            else:
                pass
    return return_dict


def process_report(client_report, abandon_group_file, new_call_details_file,
                   report_date, voice_mail_dict):
    wb = openpyxl.load_workbook('%s/%s' % (SELF_PATH, TEMPLATE_LOCATION))
    data_page = wb.get_sheet_by_name('data')
    report_page = wb.get_sheet_by_name('REPORT')
    report_page_rows = make_row_dictionary(report_page)
    data_page_rows = make_row_dictionary(data_page)
    write_outfile(client_report,
                  abandon_group_file,
                  new_call_details_file,
                  report_date,
                  voice_mail_dict,
                  report_page_rows,
                  data_page_rows,
                  wb,
                  data_page,
                  report_page)


def get_client_dictionary():
    wb = openpyxl.load_workbook(SELF_PATH + '\\bin\\CONFIG.xlsx')
    ws = wb.get_sheet_by_name("CLIENT_DICT")
    return_dict = defaultdict(list)
    for row_number in range(1, ws.max_row + 1):
        client_name = valid_input('A', row_number, ws, 'S')
        client_number = valid_input('B', row_number, ws, 'N')
        return_dict[client_number] = client_name
    return return_dict


def clear_used_data(folder_to_clear):
    folder = (SELF_PATH + folder_to_clear)
    for the_file in os.listdir(folder):
        file_path = os.path.join(folder, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            print(e)


def clear_eca_gen_tab(eca_gen_sheet):
    for row_position in range(5, eca_gen_sheet.max_row + 1):
        eca_gen_sheet['A%d' % row_position] = ""
        eca_gen_sheet['B%d' % row_position] = ""
        eca_gen_sheet['C%d' % row_position] = ""
        eca_gen_sheet['D%d' % row_position] = ""
        eca_gen_sheet['E%d' % row_position] = ""
        eca_gen_sheet['F%d' % row_position] = ""
        eca_gen_sheet['G%d' % row_position] = ""
        eca_gen_sheet['H%d' % row_position] = ""
        eca_gen_sheet['I%d' % row_position] = ""
        eca_gen_sheet['J%d' % row_position] = ""
        eca_gen_sheet['K%d' % row_position] = ""
    return eca_gen_sheet


def open_abandon_grp_report():
    raw_abandon_group_file = openpyxl.load_workbook(SELF_PATH + '\\raw\\Group Abandoned Calls.xlsx')
    page_names = raw_abandon_group_file.get_sheet_names()
    all_call_id = []

    for sheet in page_names:
        ws = raw_abandon_group_file.get_sheet_by_name(sheet)
        for row_position in range(5, ws.max_row + 1):
            if sheet == SUMMARY_PAGE:
                pass
            else:
                call_id = ws['A%d' % row_position].value
                internal_party = ws['E%d' % row_position].value
                external_party = ws['F%d' % row_position].value
                answered = ws['G%d' % row_position].value
                start_time = ws['I%d' % row_position].value
                end_time = ws['J%d' % row_position].value
                call_duration = ws['K%d' % row_position].value
                individual_call_id = AbandonGrp()
                individual_call_id.set_call(call_id,
                                            internal_party,
                                            external_party,
                                            answered,
                                            start_time,
                                            end_time,
                                            call_duration)
                if individual_call_id not in all_call_id:
                    all_call_id.append(individual_call_id)
                all_call_id.sort(key=lambda x: x.get_call_id(), reverse=False)

    eca_gen_sheet = raw_abandon_group_file.get_sheet_by_name(ABANDON_CALLS_GROUP_PAGE_NAME)
    clear_eca_gen_tab(eca_gen_sheet)

    row_position = 5
    for call in all_call_id:
        eca_gen_sheet['A%d' % row_position] = call.get_call_id()
        eca_gen_sheet['E%d' % row_position] = call.get_internal_party()
        eca_gen_sheet['F%d' % row_position] = call.get_external_party()
        eca_gen_sheet['G%d' % row_position] = call.get_bool_answered()
        eca_gen_sheet['I%d' % row_position] = call.get_start_time()
        eca_gen_sheet['J%d' % row_position] = call.get_end_time()
        eca_gen_sheet['K%d' % row_position] = call.get_call_duration()
        row_position += 1
    return raw_abandon_group_file


def process_duplicates(duplicate_dictionary_calls_to_remove, client_report):
    for client_in_report in client_report:
        for client_number_in_duplicate_dictionary in duplicate_dictionary_calls_to_remove:
            number_of_duplicates_by_client = duplicate_dictionary_calls_to_remove[client_number_in_duplicate_dictionary]
            if client_in_report.name == client_number_in_duplicate_dictionary:
                client_in_report.set_duplicates(number_of_duplicates_by_client)
                client_in_report.set_remove_lost_calls(number_of_duplicates_by_client)
            else:
                pass


def get_sec(time_string):
    try:
        h, m, s = [int(float(i)) for i in time_string.split(':')]
    except TypeError:
        return 0
    return convert_sec(h, m, s)


def convert_sec(h, m, s):
    return (3600 * int(h)) + (60 * int(m)) + int(s)


def convert_time_stamp(convert_seconds):
    minutes, seconds = divmod(convert_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return "%d:%02d:%02d" % (hours, minutes, seconds)


def binary_search(call_details_file, call_id):
    ws = call_details_file.get_sheet_by_name(CALL_DETAILS_FIRST_PAGE_NAME)
    first = 3
    last = ws.max_row
    found = False
    midpoint = 0

    while first <= last and not found:
        midpoint = (first + last) // 2
        call_id_number = call_id[9:]
        id_to_inspect = ws['A%d' % midpoint].value[9:]
        if id_to_inspect == call_id_number:
            found = True
        else:
            if call_id_number < id_to_inspect:
                last = midpoint - 1
            else:
                first = midpoint + 1
    return found, midpoint


def show_contents(client_report):
    for client in client_report:
        client.__str__()


def show_ticker(client_report):
    for client in client_report:
        client.print_ticker()


def prepare_raw_call_details(ws, call_details_file, cradle_grave_call_id_list):
    for calc_m_row_position in range(4, ws.max_row + 1):
        try:
            c_value = valid_input('C', calc_m_row_position, ws, 'N')
            j_value = valid_input('J', calc_m_row_position, ws, 'N')
            ws['M%d' % calc_m_row_position] = convert_time_stamp(c_value - j_value)
        except ValueError:
            pass

    for call_id, index in enumerate(cradle_grave_call_id_list):
        call_id_from_cradle = cradle_grave_call_id_list[call_id]
        call_hold_time = call_id_from_cradle.get_total_hold_time()
        call_id_name = call_id_from_cradle.get_name()
        if call_hold_time > 0:
            found, position = binary_search(call_details_file, call_id_name)
            if found is True:
                n_value = call_hold_time
                m_value = valid_input('M', position, ws, 'N')
                ws['N%d' % position] = convert_time_stamp(n_value)
                ws['M%d' % position] = convert_time_stamp(m_value - n_value)
    call_details_file.save(SELF_PATH + '\\raw\\Call Details Modified.xlsx')


def check_date(wb, date, page_name):
    month_dict = {'Jan': '1', 'Feb': '2', 'Mar': '3', 'Apr': '4', 'May': '5', 'Jun': '6',
                  'Jul': '7', 'Aug': '8', 'Sep': '9', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
    ws = wb.active
    try:
        if page_name == ABANDON_CALLS_GROUP_PAGE_NAME:
            page_date = ws['A3'].value[5:17].replace(",", "").split(' ')
            page_date[0] = month_dict[page_date[0]]
        elif page_name == CALL_DETAILS_FIRST_PAGE_NAME:
            page_date = ws['A2'].value[5:17].replace(",", "").split(' ')
            page_date[0] = month_dict[page_date[0]]
        elif page_name == 'Cradle_to_Grave':
            page_date = ws['A3'].value[5:17].replace(",", "").split(' ')
            page_date[0] = month_dict[page_date[0]]
        elif page_name == 'hunt':
            page_date = ws['B4'].value[5:13].replace(" ", "").split('/')
        else:
            raise ValueError("Couldn't find a date to look at.")
        try:
            page_date.remove("")
        except ValueError:
            pass
        for index, month_day_year in enumerate(page_date):
            if int(month_day_year) < 10:
                page_date[index] = ('0%s' % month_day_year)

        if len(page_date[0] + page_date[1] + page_date[2]) < 7:
            inspect_date = page_date[0] + page_date[1] + '20' + page_date[2]
        else:
            inspect_date = page_date[0] + page_date[1] + page_date[2]

        if inspect_date == date:
            pass
        else:
            raise ValueError("Sources files are not the correct date. Look at page/date: %s/%s" % (page_name,
                                                                                                   inspect_date))
    except Exception as e:
        raise ValueError(e)


def initialize_reports(report_date_datetime):
    print("Ensure files are placed in their proper directory before proceeding.")
    while True:
        try:
            date_string = report_date_datetime.strftime("%m%d%Y")
            print("Attempting to to load data files from the directory.")
            client_list_file = openpyxl.load_workbook(SELF_PATH + '\\bin\\CONFIG.xlsx')
            client_list_info = client_list_file.get_sheet_by_name("CLIENT LIST INFO")
            call_details_file = openpyxl.load_workbook(SELF_PATH + '\\raw\\Call Details.xlsx')
            check_date(call_details_file, date_string, CALL_DETAILS_FIRST_PAGE_NAME)
            abandon_group_file = open_abandon_grp_report()
            check_date(abandon_group_file, date_string, ABANDON_CALLS_GROUP_PAGE_NAME)
            chronicall_report = openpyxl.load_workbook(SELF_PATH + '\\raw\\Cradle to Grave.xlsx')
            check_date(chronicall_report, date_string, 'Cradle_to_Grave')
            abandon_group_file.get_sheet_by_name("ECA GEN")
        except ValueError:
            raise ValueError("**There was an issue with loading one of the files in Add_CDetails_AbandonGRP_C2Grave**")
        else:
            print("Data files successfully loaded.")
            return client_list_info, call_details_file, abandon_group_file, chronicall_report


def custom_input(input_string, return_type, value_to_find='X'):
    while True:
        if return_type == 'C':
            return input(input_string)
        elif return_type == 'S':
            response = input(input_string)
            if (response.isalpha()) & (response[:1].upper() == value_to_find):
                return True
            return False
        elif return_type == 'B':
            if not input(input_string):
                return True
            else:
                return False
        else:
            pass
