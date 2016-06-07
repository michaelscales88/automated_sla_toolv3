import openpyxl
import sys, traceback
import os, shutil
from abandongrp import AbandonGrp
from os import path

Abandon_Calls_Group_Page = 'ECA GEN'
Summary_Page = "Summary"
date = "TEST"

def clear_ECA_GEN(eca_gen_sheet):
    for row_position in range(5, eca_gen_sheet.max_row + 1):
        eca_gen_sheet['A' + str(row_position)] = ""
        eca_gen_sheet['B' + str(row_position)] = ""
        eca_gen_sheet['C' + str(row_position)] = ""
        eca_gen_sheet['D' + str(row_position)] = ""
        eca_gen_sheet['E' + str(row_position)] = ""
        eca_gen_sheet['F' + str(row_position)] = ""
        eca_gen_sheet['G' + str(row_position)] = ""
        eca_gen_sheet['H' + str(row_position)] = ""
        eca_gen_sheet['I' + str(row_position)] = ""
        eca_gen_sheet['J' + str(row_position)] = ""
        eca_gen_sheet['K' + str(row_position)] = ""
    return eca_gen_sheet
    
def open_abandonGrp():
    raw_abandon_group_file = openpyxl.load_workbook(os.path.dirname(path.dirname(path.abspath(__file__))) + '/Add_CDetails_AbandonGRP_C2Grave/Group Abandoned Calls.xlsx') 
    page_names = raw_abandon_group_file.get_sheet_names()
    all_call_ID = []
 
    for sheet in page_names:
        ws = raw_abandon_group_file.get_sheet_by_name(sheet)
        for row_position in range(5, ws.max_row):
            if(sheet == Summary_Page):
                pass
            else:
                Call_ID = ws['A' + str(row_position)].value
                Internal_Party = ws['E' + str(row_position)].value
                External_Party = ws['F' + str(row_position)].value
                Answered = ws['G' + str(row_position)].value
                Start_Time = ws['I' + str(row_position)].value
                End_Time = ws['J' + str(row_position)].value
                Call_Duration = ws['K' + str(row_position)].value

                individual_call_ID = AbandonGrp()
                individual_call_ID.setRow(Call_ID, Internal_Party, External_Party, Answered, Start_Time, End_Time, Call_Duration)
                if individual_call_ID not in all_call_ID:
                    all_call_ID.append(individual_call_ID)
                all_call_ID.sort(key=lambda x: x.Call_ID, reverse=False)

    eca_gen_sheet = raw_abandon_group_file.get_sheet_by_name(Abandon_Calls_Group_Page)
    clear_ECA_GEN(eca_gen_sheet)
    
    row_position = 5
    for call in all_call_ID:
        eca_gen_sheet['A' + str(row_position)] = call.Call_ID
        eca_gen_sheet['E' + str(row_position)] = call.Internal_Party
        eca_gen_sheet['F' + str(row_position)] = call.External_Party
        eca_gen_sheet['G' + str(row_position)] = call.Answered
        eca_gen_sheet['I' + str(row_position)] = call.Start_Time
        eca_gen_sheet['J' + str(row_position)] = call.End_Time
        eca_gen_sheet['K' + str(row_position)] = call.Call_Duration
        row_position += 1

##    raw_abandon_group_file.save(os.path.dirname(path.dirname(path.abspath(__file__))) + '/hunt_group_dir/' + date + 'Group Abandoned Calls.xlsx')
    for call in all_call_ID:
        call.__str__()

open_abandonGrp()
